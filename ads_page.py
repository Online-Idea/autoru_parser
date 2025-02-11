import logging
import sys
import time

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from pandas import DataFrame

from utils.auth_autoru import check_auth_page
from utils.browser_driver import browser_driver
from utils.captcha import check_captcha
from utils.email_sender import send_email_to_client
from utils.parse_autoru import parse_autoru_mark
from utils.parse_avito import parse_avito
from utils.random_wait import random_wait
from utils.result_processing import format_work, dealer_data, dealers_pandas, final_file_path

# import twocaptcha

start = time.perf_counter()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y.%m.%d %H:%M:%S",
)

driver = browser_driver()

if len(sys.argv) > 1 and sys.argv[1] == '--by_request':  # Одинарный запуск
    marks = pd.read_excel('start.xlsx', sheet_name='По марке (по запросу)')
else:  # Для автоматического запуска
    marks = pd.read_excel('start.xlsx', sheet_name='По марке (автоматом)')

# Сортирую по сайту, региону и марке чтобы не парсить одно и то же для разных клиентов
marks = marks.sort_values(by=['Сайт', 'Регион', 'Марка'])

# Считаю по почте. Это нужно для того чтобы в дальнейшем отправлять одно письмо с несколькими файлами
# вместо нескольких писем с одним файлом
emails_count = marks[marks['Активно'] == 'да']['Почты клиентов'].value_counts()
# Для каждого значения почт создаю список и добавляю его в словарь
emails_files = {i: [] for i in emails_count.index}

previous_settings = ''
previous_df: DataFrame
previous_final_file = ''

for _, mark in marks.iterrows():
    # Только активные
    if str(mark['Активно']).lower() != 'да':
        continue

    # Настройки парсера
    logging.info(f'\n{mark}')
    client = mark['Наш клиент']
    autoru_name = mark['Имя клиента на авто.ру']
    mark_name = mark['Марка']
    site = mark['Сайт'].lower()
    mark_url = mark['Ссылка']
    region = mark['Регион']
    client_email = mark['Почты клиентов']

    # Путь к готовому файлу
    final_file = final_file_path(client, site)

    # Если текущие настройки по Сайту, Региону, Марке и Ссылке не равны прошлым настройкам значит парсим
    current_settings = f'{site} {region} {mark_name} {mark_url}'
    if current_settings != previous_settings:
        df = pd.DataFrame({'mark_model': [], 'complectation': [], 'modification': [], 'year': [], 'dealer': [],
                           'price_with_discount': [], 'price_no_discount': [], 'with_nds': [], 'link': [],
                           'condition': [], 'in_stock': [], 'services': [], 'tags': [], 'photos': []})

        cars = None
        if site in ['авто.ру', 'автору']:
            cars = parse_autoru_mark(mark_url, driver, region)
        elif site == 'авито':
            cars = parse_avito(mark_url, driver, mark_name)

        if cars:
            df = df._append(dealer_data(client, cars, final_file, region))
        else:  # Нет объявлений
            continue

        # Обрабатываю собранные данные
        file_after_pandas = dealers_pandas(df, autoru_name)
        random_wait()
        if site in ['авто.ру', 'автору']:
            check_captcha(driver)
            check_auth_page(driver)

    # Иначе берём уже спарсенные объявления
    else:
        # Беру выдачу с прошлого файла и сохраняю в новом
        previous_workbook = load_workbook(previous_final_file)
        previous_ads_sheet = previous_workbook['Выдача']
        current_workbook = Workbook()
        current_ads_sheet = current_workbook.active
        current_ads_sheet.title = 'Выдача'
        for row in previous_ads_sheet.iter_rows(values_only=True):
            current_ads_sheet.append(row)
        current_workbook.save(final_file)

        file_after_pandas = dealers_pandas(previous_df, autoru_name)

    # Форматирую в читабельный вид
    format_work(file_after_pandas, autoru_name, final_file)

    # Отправляю письмо с готовыми сравнительными
    if str(client_email) != 'nan':
        emails_files[client_email].append(final_file)

        if len(emails_files[client_email]) == emails_count[client_email]:
            send_email_to_client(client_email, emails_files[client_email])

    previous_settings = current_settings
    previous_df = df
    previous_final_file = final_file

driver.quit()

logging.info(f'Общее время парсинга: {time.perf_counter() - start:.3f}')
