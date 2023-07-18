import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn
from pandas import DataFrame

# Имена столбцов по-русски
COLUMNS_RUS = {
    'mark_model': 'Марка, модель',
    'complectation': 'Комплектация',
    'modification': 'Модификация',
    'year': 'Год',
    'dealer': 'Имя дилера',
    'min_price_with_discount': 'Мин. цена со скидками',
    'min_price_no_discount': 'Мин. цена без скидок',
    'max_price': 'Макс. цена',
    'min_price_with_discount_difference': 'Разница мин. цена со скидками',
    'min_price_no_discount_difference': 'Разница мин. цена без скидок',
    'max_price_difference': 'Разница макс. цена',
    'stock': 'В наличии',
    'for_order': 'Под заказ',
    'link': 'Ссылка',
    'price_with_discount': 'Цена со скидками',
    'price_no_discount': 'Цена без скидок',
    'with_nds': 'Цена с НДС',
    'condition': 'Состояние',
    'in_stock': 'Наличие',
    'services': 'Услуги',
    'tags': 'Стикеры',
    'photos': 'Количество фото',
    'position_actual': 'Позиция по актуальности',
    'position_total': 'Позиция общая',
}

# Цвета заливок
BLUE_FILL = PatternFill(start_color='DEE6EF', end_color='DEE6EF', fill_type='solid')
RED_FILL = PatternFill(start_color='E0C2CD', end_color='E0C2CD', fill_type='solid')
GREEN_FILL = PatternFill(start_color='DDE8CB', end_color='DDE8CB', fill_type='solid')
FONT = Font(name="Calibri", bold=True)


def final_file_path(client: str, site: str) -> str:
    """
    Файл в котором будут все данные
    @param client: имя нашего клиента
    @param site: сайт который парсим
    @return: путь к готовому файлу
    """
    today = datetime.now().strftime('%d.%m.%Y')
    file_name = f'Сравнение {client} ({site}) {today}.xlsx'
    return os.path.join('results', file_name)


def dealer_data(client: str, cars: list[dict], final_file: str) -> DataFrame:
    """
    Обработка данных объявления
    @param client: имя нашего клиента
    @param cars: лист словарей с данными автомобилей
    @param final_file: путь к готовому файлу
    @return: Pandas DataFrame с автомобилями дилера
    """
    df = pd.DataFrame.from_records(cars)

    # Заполняю пустые комплектации для расчётов
    df['complectation'] = df['complectation'].fillna('empty')

    # Позиция общая
    df['position_total'] = df.reset_index().index + 1

    # Позиция по актуальности
    df['position_actual'] = df.groupby(['mark_model', 'complectation', 'modification', 'year']).cumcount() + 1

    # Сортирую по актуальности
    df = df.sort_values(by=['mark_model', 'complectation', 'modification', 'year', 'position_actual'])

    # Чищу цены
    df['price_with_discount'] = df['price_with_discount'].str.replace(r'\D+', '', regex=True)
    df['price_no_discount'] = df['price_no_discount'].str.replace(r'\D+', '', regex=True)

    # Меняю типы
    df[['year', 'price_with_discount', 'price_no_discount']] = \
        df[['year', 'price_with_discount', 'price_no_discount']] \
            .apply(lambda x: pd.to_numeric(x, errors='coerce'))

    # Меняю столбцы местами
    df = df[['mark_model', 'complectation', 'modification', 'year', 'dealer',
             'price_with_discount', 'price_no_discount', 'with_nds',
             'position_actual', 'position_total',
             'link', 'condition', 'in_stock', 'services', 'tags', 'photos']]

    # Сохраняю выдачу
    with pd.ExcelWriter(final_file, engine='xlsxwriter') as writer:
        df.T.reset_index().T.to_excel(writer, sheet_name='Выдача', header=False, index=False)

    return process_raw_ads(df)


def process_raw_ads(df: DataFrame) -> DataFrame:
    """
    # Обрабатываю выдачу
    @param df: pandas DataFrame с сырыми объявлениями
    @return: pandas DataFrame с обработанными объявлениями
    """
    # Сортирую
    df = df.sort_values(by=['mark_model', 'complectation', 'modification', 'year', 'price_with_discount'])

    # Количество автомобилей
    df_in_stock = df[df['in_stock'] == 'В наличии']
    count_df = df_in_stock.groupby(['mark_model', 'complectation', 'modification', 'year', 'dealer']) \
        .size().reset_index(name='stock')
    df = pd.merge(df, count_df, on=['mark_model', 'complectation', 'modification', 'year', 'dealer'], how='left')

    df_for_order = df[df['in_stock'].isin(['В пути', 'На заказ'])]
    count_df = df_for_order.groupby(['mark_model', 'complectation', 'modification', 'year', 'dealer']) \
        .size().reset_index(name='for_order')
    df = pd.merge(df, count_df, on=['mark_model', 'complectation', 'modification', 'year', 'dealer'], how='left')

    # Удаляю дубликаты
    # df = df.drop_duplicates(subset=['mark_model', 'complectation', 'modification', 'year'])
    df = df.drop_duplicates(subset=['mark_model', 'complectation', 'modification', 'year', 'dealer'])

    # Позиция по цене
    df['position_price'] = df.groupby(['mark_model', 'complectation', 'modification', 'year']).cumcount() + 1

    # Убираю больше ненужное 'empty' из комплектаций
    df['complectation'] = df['complectation'].str.replace('empty', '')

    return df


def dealers_pandas(df: DataFrame, autoru_name: str) -> str:
    """
    Здесь обработка данных, собранных со страниц дилеров
    @param df: Pandas DataFrame с данными дилеров
    @param autoru_name: имя клиента на авто.ру
    @return: имя готового файла
    """
    # Удаляю ненужные столбцы
    df = df.drop(columns=['condition', 'in_stock'])

    # Считаю разницу
    lookup_df = df[df['dealer'] == autoru_name]
    merged_df = pd.merge(df, lookup_df, on=['mark_model', 'complectation', 'modification', 'year'],
                         suffixes=('_data', '_lookup'), how='left')
    merged_df['price_with_discount_difference'] = merged_df['price_with_discount_data'] - merged_df[
        'price_with_discount_lookup']
    merged_df['price_no_discount_difference'] = merged_df['price_no_discount_data'] - merged_df[
        'price_no_discount_lookup']

    # Удаляю столбцы по которым считал разницы
    merged_df = merged_df.filter(regex='^(?!.*_lookup)')

    # Убираю более ненужные суффиксы
    merged_df.columns = merged_df.columns.str.replace('_data', '')

    # Сортировка
    merged_df = merged_df.sort_values(
        by=['mark_model', 'complectation', 'modification', 'year', 'price_with_discount'])

    # Меняю столбцы местами
    merged_df = merged_df[['mark_model', 'complectation', 'modification', 'year', 'dealer',
                           'price_with_discount', 'price_no_discount',
                           'price_with_discount_difference', 'price_no_discount_difference',
                           'position_price', 'position_actual', 'link', 'stock', 'for_order']]

    merged_df = merged_df.fillna('')

    # Перевожу имена столбцов
    merged_df = merged_df.rename(columns={
        'mark_model': 'Марка, модель',
        'complectation': 'Комплектация',
        'modification': 'Модификация',
        'year': 'Год',
        'dealer': 'Имя дилера',
        'price_with_discount': 'Цена со скидками',
        'price_no_discount': 'Цена без скидок',
        'price_with_discount_difference': 'Разница цены со скидками',
        'price_no_discount_difference': 'Разница цены без скидок',
        'position_price': 'Позиция по цене',
        'position_actual': 'Позиция по актуальности',
        'stock': 'В наличии',
        'for_order': 'Под заказ',
        'link': 'Ссылка',
    })

    # Это для второго листа на котором не очищаю повторы
    full_df = merged_df.copy()

    # Очищаю повторы данных автомобиля чтобы лучше смотрелось
    is_duplicate = merged_df.duplicated(subset=['Марка, модель', 'Комплектация', 'Модификация', 'Год'], keep='first')
    merged_df.loc[is_duplicate, 'Комплектация'] = ''
    merged_df.loc[is_duplicate, 'Модификация'] = ''
    merged_df.loc[is_duplicate, 'Год'] = ''
    merged_df.loc[merged_df['Марка, модель'].duplicated(), 'Марка, модель'] = ''

    # Сохраняю
    file_name = 'after_pandas.xlsx'
    with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:
        merged_df.T.reset_index().T.to_excel(writer, sheet_name='Первый вариант', header=False, index=False)
        full_df.T.reset_index().T.to_excel(writer, sheet_name='Второй вариант', header=False, index=False)

    return file_name


def data_work(cars, dealer):
    # Обрабатываю результат
    df = pd.DataFrame.from_records(cars)

    # Меняю типы
    df[['year', 'min_price_with_discount', 'min_price_no_discount', 'max_price', 'stock', 'for_order']] = \
        df[['year', 'min_price_with_discount', 'min_price_no_discount', 'max_price', 'stock', 'for_order']] \
            .apply(pd.to_numeric)

    # Удаляю Рынок т.к. самый дешёвый будет виден по сортировке
    index_to_drop = df[df['dealer'] == 'Рынок'].index
    df = df.drop(index_to_drop)

    # Этот dataframe используется для столбцов разницы
    lookup_df = df[df['dealer'] == dealer]
    merged_df = pd.merge(df, lookup_df, on=['mark_model', 'complectation', 'modification', 'year'],
                         suffixes=('_data', '_lookup'))
    merged_df['min_price_with_discount_difference'] = merged_df['min_price_with_discount_data'] - merged_df[
        'min_price_with_discount_lookup']
    merged_df['min_price_no_discount_difference'] = merged_df['min_price_no_discount_data'] - merged_df[
        'min_price_no_discount_lookup']
    merged_df['max_price_difference'] = merged_df['max_price_data'] - merged_df['max_price_lookup']

    # Сортировка
    merged_df = merged_df.sort_values(
        by=['mark_model', 'complectation', 'modification', 'year', 'min_price_with_discount_data'])

    # Удаляю столбцы по которым считал разницы
    merged_df = merged_df.filter(regex='^(?!.*_lookup)')

    merged_df.columns = merged_df.columns.str.replace('_data', '')

    # Меняю столбцы местами
    merged_df = merged_df[['mark_model', 'complectation', 'modification', 'year', 'dealer',
                           'min_price_with_discount', 'min_price_no_discount', 'max_price',
                           'min_price_with_discount_difference', 'min_price_no_discount_difference',
                           'max_price_difference', 'stock', 'for_order', 'link']]

    merged_df = merged_df.fillna('')

    # Удаляю строки с пустыми ценами
    merged_df = merged_df.drop(merged_df[(merged_df['max_price'] == '') &
                                         (merged_df['max_price_difference'] == '')].index)

    # Очищаю повторы данных автомобиля чтобы лучше смотрелось
    merged_df.loc[merged_df['mark_model'].duplicated(), 'mark_model'] = ''
    is_duplicate = merged_df.duplicated(subset=['complectation', 'modification', 'year'], keep='first')
    merged_df.loc[is_duplicate, 'complectation'] = ''
    merged_df.loc[is_duplicate, 'modification'] = ''
    merged_df.loc[is_duplicate, 'year'] = ''

    # Перевожу имена столбцов
    merged_df = merged_df.rename(columns=COLUMNS_RUS)

    file_name = 'after_pandas.xlsx'
    merged_df.T.reset_index().T.to_excel(file_name, sheet_name='Все', header=False, index=False)
    return file_name


def format_work(xlsx_file: str, autoru_name: str, final_file: str) -> None:
    """
    Форматирование готового файла в читабельный вид
    @param xlsx_file: xlsx файл после dealers_pandas
    @param autoru_name: имя клиента на авто.ру
    @param final_file: путь к готовому файлу
    """
    # Открываю файл после dealers_pandas
    after_pandas_wb = load_workbook(xlsx_file)

    # Открываю финальный файл
    final_file_wb = load_workbook(final_file)

    # Перевожу имена столбцов на листе Выдача
    ads_sheet = final_file_wb['Выдача']
    for col in range(1, ads_sheet.max_column + 1):
        cell = ads_sheet.cell(row=1, column=col)
        if cell.value in COLUMNS_RUS:
            cell.value = COLUMNS_RUS[cell.value]

    sheets = ['Первый вариант', 'Второй вариант']
    # Копирую листы
    for sheet in sheets:
        after_pandas_ws = after_pandas_wb[sheet]
        final_file_ws = final_file_wb.create_sheet(sheet)
        for row in after_pandas_ws.iter_rows():
            for cell in row:
                final_file_ws[cell.coordinate].value = cell.value

    # Добавляю лист Выдача к списку листов и форматирую все листы
    sheets.append('Выдача')
    for sheet in sheets:
        final_file_ws = final_file_wb[sheet]

        # Буквы нужных столбцов
        car_params = []
        price_cols = []
        difference_cols = []
        stock_cols = []
        position_actual_col = []
        for col in range(1, final_file_ws.max_column + 1):
            cell = final_file_ws.cell(row=1, column=col).value
            col_letter = get_column_letter(col)
            if cell in ['Марка, модель', 'Комплектация', 'Модификация', 'Год']:
                car_params.append(col)
            elif cell == 'Имя дилера':
                dealer_col = col_letter
            elif 'цен' in cell.lower():
                price_cols.append(col_letter)
            elif cell == 'Ссылка':
                link_col = col_letter
            elif cell == 'В наличии' or cell == 'Под заказ':
                stock_cols.append(col_letter)
            elif 'Позиция' in cell:
                position_actual_col.append(col_letter)

            if 'Разница' in cell:
                difference_cols.append(col_letter)

        # Границы над уникальными автомобилями
        border = Border(top=Side(style='medium', color='000000'))
        if sheet != 'Выдача':
            sheet_for_borders = final_file_wb['Второй вариант']
        else:
            sheet_for_borders = final_file_wb[sheet]
        prev_car, curr_car = '', ''
        for row in range(2, sheet_for_borders.max_row + 1):
            for car_param in car_params:
                curr_car += str(sheet_for_borders.cell(row=row, column=car_param).value)

            if curr_car != prev_car:
                for col2 in range(1, final_file_ws.max_column + 1):
                    cell = final_file_ws.cell(row=row, column=col2)
                    cell.border = border

            prev_car = curr_car
            curr_car = ''

        last_row = sheet_for_borders.max_row + 1
        for col2 in range(1, final_file_ws.max_column + 1):
            cell = final_file_ws.cell(row=last_row, column=col2)
            cell.border = border

        # Меняю ширину столбцов
        for column in final_file_ws.columns:
            final_file_ws[column[0].coordinate].alignment = Alignment(wrap_text=True)
            max_length = 0
            column_name = column[0].column_letter

            if column_name == link_col:
                final_file_ws.column_dimensions[column_name].width = 20
            elif column_name in price_cols or column_name in stock_cols or column_name in position_actual_col:
                final_file_ws.column_dimensions[column_name].width = 13
            else:
                for cell in column:
                    if cell.row == 1:
                        continue
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                final_file_ws.column_dimensions[column_name].width = adjusted_width

        # Выделяю голубым нашего дилера
        for row in range(2, final_file_ws.max_row + 1):
            if final_file_ws[dealer_col + str(row)].value == autoru_name:
                prices_range = final_file_ws[f'{dealer_col}{str(row)}:{price_cols[-1]}{str(row)}']
                for cell in prices_range[0]:
                    cell.fill = BLUE_FILL
                    # cell.font = FONT

            # Выделяю красным тех кто дешевле нас и зеленым тех кто дороже нас
            for col in difference_cols:
                cell = final_file_ws[col + str(row)]
                if cell.value:
                    if cell.value < 0:
                        cell.fill = RED_FILL
                    elif cell.value > 0:
                        cell.fill = GREEN_FILL

        # Высота первой строки в 3 строки
        final_file_ws.row_dimensions[1].height = 45

        # Выравнивание по центру первой строки и жирный шрифт
        for cell in final_file_ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = FONT

        # Закрепляю строки и столбцы
        final_file_ws.freeze_panes = 'C2'

        # Автофильтр
        final_file_ws.auto_filter.ref = final_file_ws.dimensions

    # Двигаю лист Выдачи правее Сравнительных
    final_file_wb.move_sheet('Выдача', 2)

    after_pandas_wb.close()
    final_file_wb.save(final_file)
    return
